#! /usr/bin/python3

import logging
import random
from datetime import datetime, time
from os import environ
from pathlib import Path
from time import sleep

import lyricsgenius
import openpyxl
import schedule
from pyshorteners import Shortener, Shorteners
from pytz import timezone, utc
from twilio.rest import Client


def run_schedule(lower_bound, upper_bound):
    """
    Executes scheduled jobs when jobs are available. If an exception is
    encountered, the details are logged and the job queue is cancelled and
    restarted (to prevent constant retries)

    :param int lower_bound: lower bound of interval (in minutes)
    :param int upper_bound: upper bound of interval (in minutes)
    """
    schedule_job(lower_bound, upper_bound)

    while True:
        try:
            schedule.run_pending()
        except Exception as e:
            restart_job(lower_bound=lower_bound, upper_bound=upper_bound)
            logging.exception(e)

        sleep(59)


def schedule_job(lower_bound, upper_bound):
    """
    Schedules recurring tasks at random intervals between the provided
    bounds in minutes

    :param int lower_bound: lower bound of interval (in minutes)
    :param int upper_bound: upper bound of interval (in minutes)
    """
    schedule.every(lower_bound).to(upper_bound).minutes.do(send_earworm,
                                                           sheet=ws,
                                                           genius=genius_client,
                                                           access_token=bitly_token,
                                                           twilio=twilio_client,
                                                           recipient=environ.get('RECIPIENT'))


def restart_job(lower_bound, upper_bound):
    """
    In the event of an exception occurring during the execution of a job, this
    function is called. This prevent rapid consecutive executions of the same job

    :param int lower_bound: lower bound of interval (in minutes)
    :param int upper_bound: upper bound of interval (in minutes)
    """
    schedule.clear()
    schedule_job(lower_bound=lower_bound, upper_bound=upper_bound)


def send_earworm(sheet, genius, access_token, twilio, recipient):
    """
    Using the provided earworm library, a random song is chosen. Then, the genius
    client gathers the url to the lyrics for said song. After the full link is
    retrieved, the link is shortened using bitly. Finally, the earworm message
    is created and sent to the given recipient over SMS via the twilio client

    :param Worksheet sheet: library of send_earworms
    :param Genius genius: genius client (genius.com)
    :param dict access_token: access token for bitly
    :param Client twilio: twilio client used to send message
    :param str recipient: recipient's phone number
    """
    if get_availability():
        logging.debug('Gathering earworm')
        song_artist, song_title, earworm_lyrics = get_earworm(sheet)
        original_url = get_genius_link(genius=genius,
                                       artist=song_artist,
                                       title=song_title)
        short_url = shorten_link(long_url=original_url, access_token=access_token)
        earworm_message = build_message(lyrics=earworm_lyrics, url=short_url)
        send_sms(client=twilio, message=earworm_message, recipient=recipient)
        # duplicate for testing
        send_sms(client=twilio, message=earworm_message, recipient=environ.get('MY_NUMBER'))
    else:
        logging.info(f'Skipping this job as it falls outside of the specified availability window')


def get_earworm(sheet):
    """
    Chooses random song from the library and returns the title, artist, and earworm
    
    :param Worksheet sheet: worksheet containing earworm library
    """
    row = random.randint(2, sheet.max_row)
    artist = sheet.cell(row=row, column=1).value.strip()
    title = sheet.cell(row=row, column=2).value.strip()
    earworm = sheet.cell(row=row, column=3).value.strip()

    logging.info(f'{artist} - "{title}"')
    return artist, title, earworm


def get_genius_link(genius, title, artist):
    """
    Uses Genius client to search for the song via artist and title query.
    Once a match is found, the url for the result is returned

    :param Genius genius: genius client (genius.com)
    :param str title: name of song
    :param str artist: name of artist
    """
    url = genius.search_song(title=title, artist=artist).url

    logging.debug(f'Genius URL for {artist} - {title}: {url}')
    return url


def shorten_link(long_url, access_token):
    """
    Takes a long url and shortens it using pyshorteners (this implementation
    is for bitly, but pyshorteners supports other shorteners as well)

    :param str long_url: full link to shorten
    :param dict access_token: access token for bitly
    """
    shortener = Shortener(Shorteners.BITLY, **access_token)
    short_url = shortener.short(long_url)

    logging.debug(f'{long_url} shortened to {short_url}')
    return short_url


def build_message(lyrics, url, emoji='🎶🎵🎶'):
    """
    Builds a message body by padding either side of the earworm lyrics with
    emojis and appending a link to the lyrics to the end.

    :param str lyrics: earworm (catchy lyrics)
    :param str url: url to song's lyrics on genius
    :param str emoji: emojis to pad earworm text with
    """
    earworm = f'{emoji}' \
        f'\n{lyrics}' \
        f'\n{emoji}' \
        f'\n{url}'

    logging.debug(f'Earworm message: \n"{earworm}"')
    return earworm


def send_sms(client, message, recipient):
    """
    Using the provided twilio client, the given message is transmitted
    to the recipient via SMS

    :param Client client: twilio client used to send message
    :param str message: message body
    :param str recipient: recipient's phone number
    """
    sender = environ.get('TWILIO_NUMBER')
    client.messages.create(body=message, from_=sender, to=recipient)
    logging.info(f'Earworm sent to recipient')


def get_edt_time():
    """
    Returns current time in EDT
    """
    utc_dt = utc.localize(datetime.utcnow())
    timezone_ = timezone('US/Eastern')
    edt_dt = utc_dt.astimezone(tz=timezone_)

    logging.debug(f'UTC: {utc_dt} \tEDT: {edt_dt}')
    return edt_dt


def get_availability():
    """
    Checks if current time is within availability range
    """
    start = time(hour=9, minute=0, second=0, tzinfo=timezone('US/Eastern'))
    end = time(hour=23, minute=0, second=0, tzinfo=timezone('US/Eastern'))
    is_available = start <= get_edt_time().time() <= end

    logging.debug(f'Current time is{"" if is_available else " not"} within availability range')
    return is_available


def logger_setup():
    """
    Sets up logger with specified format and explicitly converts time to EDT
    regardless of local timezone
    """
    logging.basicConfig(filename=f'{Path(__file__).stem}.log',
                        level=logging.INFO,
                        format=' %(asctime)s.%(msecs)03d - %(levelname)s - '
                               '<%(funcName)s>: %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S')
    logging.Formatter.converter = custom_time
    logging.getLogger("twilio").setLevel(logging.WARNING)
    logging.getLogger("schedule").setLevel(logging.WARNING)


def custom_time(*args):
    """
    Converts local time to custom timezone (EDT)
    This is used to convert the logging timestamps to desired timezone
    rather than defaulting to server's local time
    """
    return get_edt_time().timetuple()


def get_clients():
    """
    Initializes genius client and twilio client using environment variables
    """
    account_sid = environ.get('TWILIO_ACCOUNT_SID')
    auth_token = environ.get('TWILIO_AUTH_TOKEN')
    genius_token = environ.get('GENIUS_TOKEN')

    genius = lyricsgenius.Genius(genius_token)
    twilio = Client(account_sid, auth_token)

    logging.debug(f'Clients for Genius and Twilio created')
    return genius, twilio


if __name__ == '__main__':
    logger_setup()
    genius_client, twilio_client = get_clients()
    bitly_token = {'bitly_token': environ.get('BITLY_TOKEN')}

    wb = openpyxl.load_workbook(Path('../earworm_library/earworms.xlsx'))
    ws = wb.active

    run_schedule(lower_bound=75, upper_bound=5 * 60)
