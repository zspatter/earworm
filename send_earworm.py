import random
from os import environ
from pathlib import Path

import lyricsgenius
import openpyxl
from pyshorteners import Shortener, Shorteners
from twilio.rest import Client


def get_clients():
    account_sid = environ.get('TWILIO_ACCOUNT_SID')
    auth_token = environ.get('TWILIO_AUTH_TOKEN')
    genius_token = environ.get('GENIUS_TOKEN')

    genius = lyricsgenius.Genius(genius_token)
    twilio = Client(account_sid, auth_token)

    return genius, twilio


def get_earworm(sheet):
    row = random.randint(2, sheet.max_row)
    title = sheet.cell(row=row, column=1).value.strip()
    artist = sheet.cell(row=row, column=2).value.strip()
    earworm = sheet.cell(row=row, column=3).value.strip()

    return title, artist, earworm


def get_genius_link(genius, title, artist):
    return genius.search_song(title=title, artist=artist).url


def shorten_link(long_url, access_token):
    shortener = Shortener(Shorteners.BITLY, **access_token)

    return shortener.short(long_url)


def send_sms(client, message):
    sender = environ.get('TWILIO_NUMBER')
    recipient = environ.get('MY_NUMBER')

    client.messages.create(body=message, from_=sender, to=recipient)


def send_earworm(sheet, genius, access_token, twilio):
    song_title, song_artist, earworm_lyrics = get_earworm(sheet)
    original_url = get_genius_link(genius=genius,
                                   title=song_title,
                                   artist=song_artist)
    short_url = shorten_link(long_url=original_url,
                             access_token=access_token)
    send_sms(client=twilio,
             message=f'{earworm_lyrics}\n{short_url}')


if __name__ == '__main__':
    bitly_token = {'bitly_token': environ.get('BITLY_TOKEN')}
    genius_client, twilio_client = get_clients()

    wb = openpyxl.load_workbook(Path('./earworms/earworms.xlsx'))
    ws = wb.active
    send_earworm(sheet=ws,
                 genius=genius_client,
                 access_token=bitly_token,
                 twilio=twilio_client)
